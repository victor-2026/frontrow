#!/usr/bin/env python3
"""Functional coverage matrix: Maestro tests × features → XLSX"""
import re, glob, os
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(["pip3", "install", "openpyxl"], check=True)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

MAESTRO_DIR = "tests/maestro"
TESTIDS_FILE = "src/testIds.ts"

hdr_font = Font(bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
alt_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def parse_yaml(path):
    with open(path) as f:
        content = f.read()
    ids = set(re.findall(r"id:\s+'([^']+)'", content))
    tags = set(re.findall(r"^\s*-\s+'?(\w+)'?", content.split('---')[1] if '---' in content else '', re.MULTILINE))
    tags -= {'runFlow', 'tapOn', 'assertVisible', 'assertNotVisible', 'scrollUntilVisible',
             'extendedWaitUntil', 'swipe', 'inputText', 'openLink', 'launchApp', 'stopApp',
             'appId', 'when', 'commands', 'direction', 'from', 'duration', 'timeout'}
    desc = ""
    for line in content.split('\n'):
        if line.startswith('#'):
            desc += line.lstrip('# ').strip() + " "
    return ids, tags, desc.strip()

def parse_testids(path):
    ids = {}
    with open(path) as f:
        content = f.read()
    for m in re.findall(r"([a-zA-Z]\w+):\s+'((?:[a-zA-Z_]\w*\.)+[a-zA-Z_]\w*)'", content):
        ids[m[1]] = m[0]
    return ids

def feature_area(filepath, ids, tags):
    area_tags = {
        'onboarding': ['onboarding'],
        'auth': ['auth', 'login', 'forgotPassword', 'recovery', 'edit-profile'],
        'events': ['events', 'eventDetail', 'browse', 'sort', 'filter', 'favorites', 'follow'],
        'tickets': ['tickets', 'buy', 'tier', 'cancel', 'transfer', 'detail'],
        'billing': ['billing', 'payment', 'buy-success', 'buy-decline'],
        'debug': ['debug', 'native', 'failure'],
        'capabilities': ['capabilities', 'haptic'],
        'smoke': ['smoke'],
    }
    fname = os.path.basename(filepath).lower()
    for area, keywords in area_tags.items():
        for kw in keywords:
            if kw in fname:
                return area
    for tid in ids:
        area = tid.split('.')[0]
        if area in area_tags:
            return area
    return 'other'

# Parse all
yaml_files = sorted(glob.glob(f"{MAESTRO_DIR}/**/*.yaml", recursive=True))
defined_ids = parse_testids(TESTIDS_FILE)

rows = []
total_used = set()
for f in yaml_files:
    ids, tags, desc = parse_yaml(f)
    area = feature_area(f, ids, tags)
    rel = os.path.relpath(f, MAESTRO_DIR)
    rows.append((area, rel, ids, tags, desc))
    total_used |= ids

# Build XLSX
wb = Workbook()
ws = wb.active
ws.title = "Coverage Matrix"

headers = ["Area", "Test File", "Description", "Tags", "TestIDs Used", "IDs Count", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

areas_order = ['smoke', 'onboarding', 'auth', 'events', 'tickets', 'billing', 'debug', 'capabilities', 'other']
area_labels = {
    'smoke': '🔵 Smoke', 'onboarding': '🟣 Onboarding', 'auth': '🟡 Auth',
    'events': '🟢 Events', 'tickets': '🔴 Tickets', 'billing': '🟠 Billing',
    'debug': '⚙️ Debug', 'capabilities': '📱 Capabilities', 'other': '⬜ Other'
}

row_idx = 2
for area in areas_order:
    area_rows = [r for r in rows if r[0] == area]
    for a, rel, ids, tags, desc in area_rows:
        ids_str = ", ".join(sorted(ids)) if ids else "—"
        status = "✅" if ids else "⚠️ no assertions"
        ws.cell(row=row_idx, column=1, value=area_labels.get(area, area))
        ws.cell(row=row_idx, column=2, value=rel)
        ws.cell(row=row_idx, column=3, value=desc[:120])
        ws.cell(row=row_idx, column=4, value=", ".join(sorted(tags)) if tags else "—")
        ws.cell(row=row_idx, column=5, value=ids_str)
        ws.cell(row=row_idx, column=6, value=len(ids))
        ws.cell(row=row_idx, column=7, value=status)
        for col in range(1, 8):
            c = ws.cell(row=row_idx, column=col)
            c.border = thin_border
            c.alignment = Alignment(vertical='top', wrap_text=True)
            if row_idx % 2 == 0:
                c.fill = alt_fill
        row_idx += 1

# Column widths
widths = [16, 50, 60, 30, 60, 10, 10]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{row_idx-1}"

# Sheet 2: Uncovered testIDs
ws2 = wb.create_sheet("Uncovered IDs")
ws2.cell(row=1, column=1, value="TestID").font = hdr_font
ws2.cell(row=1, column=1).fill = hdr_fill
ws2.cell(row=1, column=1).border = thin_border
ws2.cell(row=1, column=2, value="Screen/Module").font = hdr_font
ws2.cell(row=1, column=2).fill = hdr_fill
ws2.cell(row=1, column=2).border = thin_border

uncovered = [tid for tid in sorted(defined_ids) if tid not in total_used]
for i, tid in enumerate(uncovered, 2):
    ws2.cell(row=i, column=1, value=tid).border = thin_border
    screen = tid.split('.')[0]
    ws2.cell(row=i, column=2, value=screen).border = thin_border
    if i % 2 == 0:
        ws2.cell(row=i, column=1).fill = alt_fill
        ws2.cell(row=i, column=2).fill = red_fill

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 20

# Sheet 3: Summary
ws3 = wb.create_sheet("Summary")
ws3.cell(row=1, column=1, value="Metric").font = hdr_font
ws3.cell(row=1, column=1).fill = hdr_fill
ws3.cell(row=1, column=1).border = thin_border
ws3.cell(row=1, column=2, value="Value").font = hdr_font
ws3.cell(row=1, column=2).fill = hdr_fill
ws3.cell(row=1, column=2).border = thin_border

summary = [
    ("Total YAML files", len(yaml_files)),
    ("TestIDs defined (testIds.ts)", len(defined_ids)),
    ("TestIDs used in tests", len(total_used)),
    ("TestIDs uncovered", len(uncovered)),
    ("Screen coverage %", f"{len(total_used & set(defined_ids)) * 100 // len(defined_ids) if defined_ids else 0}%"),
    ("", ""),
    ("By Area:", ""),
]

# Count by area
area_counts = defaultdict(int)
for a, _, _, _, _ in rows:
    area_counts[area_labels.get(a, a)] += 1
for area, count in sorted(area_counts.items()):
    summary.append((area, count))

for i, (k, v) in enumerate(summary, 2):
    ws3.cell(row=i, column=1, value=k).border = thin_border
    ws3.cell(row=i, column=2, value=str(v)).border = thin_border
    if k and not k.startswith("By"):
        ws3.cell(row=i, column=1).font = Font(bold=True) if ":" in k else Font()

ws3.column_dimensions['A'].width = 30
ws3.column_dimensions['B'].width = 15

output = "docs/functional-coverage.xlsx"
wb.save(output)
print(f"✅ Saved: {output}")
print(f"   Files: {len(yaml_files)}, Defined: {len(defined_ids)}, Used: {len(total_used)}, Uncovered: {len(uncovered)}")
